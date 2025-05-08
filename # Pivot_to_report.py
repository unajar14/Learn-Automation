# Pivot_to_report
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

month = 'February'

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet, min_col=min_column+1, min_row=min_row, max_col=max_column, max_row=max_row)
categories = Reference(sheet, min_col=min_column, min_row=min_row+1, max_col=max_column, max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart, "B12")
barchart.title = "Sales by Product Line"
barchart.style = 5

for i in range(min_column+1, max_column+1):  # (B, G+1)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'
    sheet = wb['Report']

sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Calibri', bold=True, size=14)

wb.save('pivot_to_report.xlsx')